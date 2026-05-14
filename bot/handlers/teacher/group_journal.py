import logging
from aiogram import Router, types, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from bot.models.education import Lesson, Attendance, StudentGroup, HomeworkSubmission
from bot.models.user import Student
from bot.keyboards.teacher import get_group_manage_kb
from aiogram.utils.keyboard import InlineKeyboardBuilder

router = Router(name="teacher_journal")
logger = logging.getLogger(__name__)

class LessonEditStates(StatesGroup):
    waiting_for_new_date = State()

@router.callback_query(F.data.startswith("t_journal:"))
async def view_group_journal(callback: types.CallbackQuery, session: AsyncSession):
    """Сводная таблица успеваемости и посещаемости группы."""
    group_id = int(callback.data.split(":")[1])
    
    # Получаем группу и учеников
    stmt = (
        select(StudentGroup)
        .where(StudentGroup.group_id == group_id)
        .options(selectinload(StudentGroup.student).selectinload(Student.user))
    )
    students = (await session.execute(stmt)).scalars().all()
    
    text = f"📖 *Учебный журнал группы ID {group_id}*\n\n"
    text += "👤 Ученик | Посещ. | Ср. балл\n"
    text += "--------------------------------\n"
    
    for link in students:
        s_id = link.student_id
        # Посещаемость
        att_stmt = select(func.count(Attendance.id)).where(Attendance.student_id == s_id, Attendance.status == "present")
        attended = await session.scalar(att_stmt) or 0
        
        # Средний балл по ДЗ
        grade_stmt = select(func.avg(HomeworkSubmission.grade)).where(HomeworkSubmission.student_id == s_id, HomeworkSubmission.status == "accepted")
        avg_grade = await session.scalar(grade_stmt) or 0
        
        text += f"{link.student.user.full_name[:12]} | {attended} зан. | {float(avg_grade):.1f}\n"

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📥 Скачать Журнал (.xlsx)", callback_data=f"t_j_excel:{group_id}")],
        [InlineKeyboardButton(text="⬅️ Опции группы", callback_data=f"t_group:{group_id}")]
    ])
    
    await callback.message.edit_text(text, parse_mode="Markdown", reply_markup=kb)

@router.callback_query(F.data.startswith("t_j_excel:"))
async def generate_group_journal_excel(callback: types.CallbackQuery, session: AsyncSession):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    group_id = int(callback.data.split(":")[1])
    
    stmt = (
        select(StudentGroup)
        .where(StudentGroup.group_id == group_id)
        .options(selectinload(StudentGroup.student).selectinload(Student.user))
    )
    students = (await session.execute(stmt)).scalars().all()
    
    if not students:
        return await callback.answer("В группе нет учеников.", show_alert=True)
        
    wb = Workbook()
    ws = wb.active
    ws.title = f"Журнал Группы {group_id}"
    
    # Headers
    headers = ["№", "ФИО Ученика", "Посещено заянтий", "Пропущено занятий", "Средний балл ДЗ"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for i, link in enumerate(students, 1):
        s_id = link.student_id
        # Посещено
        att_stmt = select(func.count(Attendance.id)).where(Attendance.student_id == s_id, Attendance.status == "present")
        attended = await session.scalar(att_stmt) or 0
        
        # Пропущено
        abs_stmt = select(func.count(Attendance.id)).where(Attendance.student_id == s_id, Attendance.status == "absent")
        absent = await session.scalar(abs_stmt) or 0
        
        # Баллы
        grade_stmt = select(func.avg(HomeworkSubmission.grade)).where(HomeworkSubmission.student_id == s_id, HomeworkSubmission.status == "accepted")
        avg_grade = await session.scalar(grade_stmt) or 0
        
        ws.append([i, link.student.user.full_name, attended, absent, f"{float(avg_grade):.1f}"])
        
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    # Save to bytes
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    from aiogram.types import BufferedInputFile
    document = BufferedInputFile(file_stream.read(), filename=f"Журнал_Группы_{group_id}.xlsx")
    
    await callback.message.answer_document(document, caption=f"📊 Журнал группы ID {group_id}")
    await callback.answer()

@router.callback_query(F.data.startswith("t_lessons_manage:"))
async def list_group_lessons_for_edit(callback: types.CallbackQuery, session: AsyncSession, state: FSMContext):
    await state.clear()
    parts = callback.data.split(":")
    group_id = int(parts[1])
    page = int(parts[2]) if len(parts) > 2 else 1
    
    stmt = select(Lesson).where(Lesson.group_id == group_id).order_by(Lesson.lesson_date.asc())
    lessons = (await session.execute(stmt)).scalars().all()
    
    from bot.utils.pagination import Paginator
    paginator = Paginator(lessons, page=page, limit=10, callback_prefix=f"t_lessons_manage:{group_id}")
    current_items = paginator.get_page_items()
    
    builder = InlineKeyboardBuilder()
    for l in current_items:
        builder.row(types.InlineKeyboardButton(text=f"✏️ {l.lesson_date.strftime('%d.%m')} - {l.topic[:15]}", callback_data=f"t_edit_lesson:{l.id}"))
    
    paginator.add_pagination_buttons(builder)
    builder.row(types.InlineKeyboardButton(text="⬅️ Назад", callback_data=f"t_group:{group_id}"))
    
    await callback.message.edit_text(
        f"📅 *Управление уроками*\nВыберите урок для переноса или редактирования (Страница {page}/{paginator.total_pages or 1}):", 
        parse_mode="Markdown", 
        reply_markup=builder.as_markup()
    )

@router.callback_query(F.data.startswith("t_edit_lesson:"))
async def start_lesson_edit(callback: types.CallbackQuery, state: FSMContext, session: AsyncSession):
    lesson_id = int(callback.data.split(":")[1])
    stmt = select(Lesson).where(Lesson.id == lesson_id)
    lesson = (await session.execute(stmt)).scalar_one()
    
    await state.update_data(lesson_id=lesson_id, group_id=lesson.group_id)
    
    builder = InlineKeyboardBuilder()
    builder.row(types.InlineKeyboardButton(text="❌ Отмена", callback_data=f"t_lessons_manage:{lesson.group_id}"))
    
    await callback.message.edit_text(
        f"✏️ *Редактирование урока*\n\nТема: {lesson.topic}\nТекущая дата: {lesson.lesson_date}\nВремя: {lesson.lesson_time}\n\n"
        f"Введите новую дату и время в формате: `ДД.ММ.ГГГГ ЧЧ:ММ`",
        parse_mode="Markdown",
        reply_markup=builder.as_markup()
    )
    await state.set_state(LessonEditStates.waiting_for_new_date)

@router.message(LessonEditStates.waiting_for_new_date)
async def process_new_lesson_date(message: types.Message, state: FSMContext, session: AsyncSession):
    from datetime import datetime
    data = await state.get_data()
    lesson_id = data['lesson_id']
    
    try:
        new_dt = datetime.strptime(message.text, "%d.%m.%Y %H:%M")
        stmt = select(Lesson).where(Lesson.id == lesson_id)
        lesson = (await session.execute(stmt)).scalar_one()
        
        lesson.lesson_date = new_dt.date()
        lesson.lesson_time = new_dt.strftime("%H:%M")
        await session.commit()
        
        await message.answer(f"✅ Урок «{lesson.topic}» перенесен на {message.text}", reply_markup=get_group_manage_kb(lesson.group_id))
        await state.clear()
    except ValueError:
        await message.answer("❌ Неверный формат. Используйте: `ДД.ММ.ГГГГ ЧЧ:ММ`")
